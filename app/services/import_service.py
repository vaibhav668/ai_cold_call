import csv
import io
import re
import uuid
import openpyxl
from datetime import datetime, timezone
from typing import List, Dict, Any, Tuple
from sqlalchemy.ext.asyncio import AsyncSession
from app.repositories.customer import CustomerRepository
from app.repositories.import_history import ImportHistoryRepository
from app.models.customer import Customer
from app.models.import_history import ImportHistory
from app.core.logging import logger
from app.core.exceptions import AppException

# Regex validators
PHONE_REGEX = re.compile(r"^\+?[1-9]\d{1,14}$")
EMAIL_REGEX = re.compile(r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$")

class CustomerImportService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.customer_repo = CustomerRepository(db)
        self.history_repo = ImportHistoryRepository(db)

    async def import_customers(
        self,
        filename: str,
        file_content: bytes,
        uploaded_by: uuid.UUID
    ) -> ImportHistory:
        """Parses CSV/Excel file, validates customer rows, inserts valid entries, and logs import history."""
        error_details: List[Dict[str, Any]] = []
        parsed_rows: List[Dict[str, Any]] = []
        
        ext = filename.split(".")[-1].lower()
        if ext == "csv":
            parsed_rows, error_details = self._parse_csv(file_content)
        elif ext in ["xlsx", "xls"]:
            parsed_rows, error_details = self._parse_excel(file_content)
        else:
            raise AppException("Unsupported file format. Only CSV and Excel (.xlsx) files are supported.", 400)
            
        total_records = len(parsed_rows) + len(error_details)
        success_count = 0
        failed_count = len(error_details)
        
        batch_phones = set()
        valid_customers: List[Customer] = []
        
        # If headers matching failed and first error occurred
        if error_details and not parsed_rows:
            status = "failed"
            history_record = ImportHistory(
                filename=filename,
                status=status,
                total_records=total_records,
                successfully_imported=0,
                failed_records=failed_count,
                error_details=error_details,
                uploaded_by=uploaded_by
            )
            await self.history_repo.create(history_record)
            await self.db.commit()
            return history_record
        
        for idx, row in enumerate(parsed_rows):
            row_num = row.get("_row_num", idx + 2)
            first_name = str(row.get("first_name", "")).strip() if row.get("first_name") is not None else ""
            last_name = str(row.get("last_name", "")).strip() if row.get("last_name") is not None else None
            phone = "".join(str(row.get("phone_number", "")).split()) if row.get("phone_number") is not None else ""
            email = str(row.get("email", "")).strip() if row.get("email") is not None else None
            
            # Extract custom metadata variables
            custom_vars = {}
            for k, v in row.items():
                if k not in ["first_name", "last_name", "phone_number", "email", "_row_num"] and v is not None:
                    custom_vars[k] = v
            
            # Row validators
            if not first_name:
                error_details.append({"row": row_num, "error": "Missing required field: first_name"})
                failed_count += 1
                continue
                
            if not phone:
                error_details.append({"row": row_num, "error": "Missing required field: phone_number"})
                failed_count += 1
                continue
                
            if not PHONE_REGEX.match(phone):
                error_details.append({"row": row_num, "error": f"Invalid E.164 phone: '{phone}'."})
                failed_count += 1
                continue
                
            if email and not EMAIL_REGEX.match(email):
                error_details.append({"row": row_num, "error": f"Invalid email format: '{email}'."})
                failed_count += 1
                continue
                
            # Duplicate check in file batch
            if phone in batch_phones:
                error_details.append({"row": row_num, "error": f"Duplicate phone number in file: '{phone}'."})
                failed_count += 1
                continue
                
            # Duplicate check in DB
            existing_db = await self.customer_repo.get_by_phone(phone)
            if existing_db:
                error_details.append({"row": row_num, "error": f"Phone number '{phone}' already exists in database."})
                failed_count += 1
                continue
                
            batch_phones.add(phone)
            customer_obj = Customer(
                first_name=first_name,
                last_name=last_name,
                phone_number=phone,
                email=email,
                custom_variables=custom_vars,
                is_active=True
            )
            valid_customers.append(customer_obj)
            success_count += 1
            
        # Bulk creations
        if valid_customers:
            for customer in valid_customers:
                await self.customer_repo.create(customer)
                
        # Resolve status
        status = "success"
        if failed_count > 0:
            status = "partial" if success_count > 0 else "failed"
            
        history_record = ImportHistory(
            filename=filename,
            status=status,
            total_records=total_records,
            successfully_imported=success_count,
            failed_records=failed_count,
            error_details=error_details,
            uploaded_by=uploaded_by
        )
        
        await self.history_repo.create(history_record)
        await self.db.commit()
        
        return history_record

    def _parse_csv(self, file_content: bytes) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """Decode and read CSV files into dictionaries."""
        rows: List[Dict[str, Any]] = []
        errors: List[Dict[str, Any]] = []
        
        try:
            content_str = file_content.decode("utf-8")
        except UnicodeDecodeError:
            try:
                content_str = file_content.decode("latin-1")
            except Exception as e:
                errors.append({"row": 1, "error": f"Failed to decode file: {e}"})
                return [], errors
                
        reader = csv.DictReader(io.StringIO(content_str))
        
        headers = reader.fieldnames or []
        required = ["first_name", "phone_number"]
        missing = [req for req in required if req not in headers]
        if missing:
            errors.append({"row": 1, "error": f"Missing required headers: {', '.join(missing)}"})
            return [], errors
            
        for idx, row in enumerate(reader):
            row_num = idx + 2
            row_dict = dict(row)
            row_dict["_row_num"] = row_num
            rows.append(row_dict)
            
        return rows, errors

    def _parse_excel(self, file_content: bytes) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """Read Excel spreadsheets using openpyxl in read-only mode."""
        rows: List[Dict[str, Any]] = []
        errors: List[Dict[str, Any]] = []
        
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
            sheet = wb.active
            if sheet is None:
                errors.append({"row": 1, "error": "Excel sheet is empty."})
                return [], errors
                
            header_row = next(sheet.iter_rows(max_row=1, values_only=True), None)
            if not header_row:
                errors.append({"row": 1, "error": "Excel file is empty."})
                return [], errors
                
            headers = [str(cell).strip() if cell is not None else f"col_{col_idx}" for col_idx, cell in enumerate(header_row)]
            required = ["first_name", "phone_number"]
            missing = [req for req in required if req not in headers]
            if missing:
                errors.append({"row": 1, "error": f"Missing required Excel headers: {', '.join(missing)}"})
                wb.close()
                return [], errors
                
            for idx, r in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                row_num = idx + 2
                
                if all(cell is None for cell in r):
                    continue
                    
                row_dict = {}
                for col_idx, cell_value in enumerate(r):
                    if col_idx < len(headers):
                        if headers[col_idx] == "phone_number" and isinstance(cell_value, (int, float)):
                            row_dict[headers[col_idx]] = str(int(cell_value))
                        else:
                            row_dict[headers[col_idx]] = cell_value
                            
                row_dict["_row_num"] = row_num
                rows.append(row_dict)
                
            wb.close()
        except Exception as e:
            errors.append({"row": 1, "error": f"Failed to parse Excel: {str(e)}"})
            
        return rows, errors
